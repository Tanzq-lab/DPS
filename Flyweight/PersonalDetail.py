import json
import os
import pickle
import shelve
import openpyxl as openpyxl
from docx import Document


class PersonalDetail:
    """
    该类主要用于存储每个人的数据
    """

    def __init__(self, filename=""):
        # 初始化数据
        self.name = "none"
        self.address = "none"
        self.phone = "none"
        # 序列化的方式
        self.Type = "none"

        # print(filename)
        if os.path.exists(filename):
            # print("文件存在！")
            self.Type = filename.split('.')[1]
            self.decoding(filename)
        # elif os.path.exists(filename + '.dat'):
        #     self.Type = "Tanzq"
        #     self.decoding(filename)

    def clear(self):
        # 清空数据
        self.address = "none"
        self.phone = "none"

    def encoding(self):
        # 按照不同类型进行序列化操作
        if self.Type == 'json':
            print("正在进行json序列化操作...")
            filename = "Data\\" + str(self.name) + "_json.json"
            with open(filename, 'w') as fp:
                json.dump(self.__dict__, fp, ensure_ascii=False, indent=4)

        elif self.Type == 'txt':
            print("正在进行pickle序列化操作...")
            filename = "Data\\" + str(self.name) + "_pickle.txt"
            with open(filename, 'wb') as fp:
                pickle.dump(self.__dict__, fp)

        elif self.Type == 'Tanzq':
            print("正在进行shelve序列化操作...")
            filename = "Data\\" + str(self.name) + "_shelve.Tanzq"
            with shelve.open(filename) as fp:
                fp['name'] = self.name
                fp['address'] = self.address
                fp['phone'] = self.phone

        print("序列化操作完成！\n")

    def decoding(self, filename):
        # 按照不同的操作进行反序列化操作
        if self.Type == 'json':
            print("正在对{filename}进行json反序列化...".format(filename=filename))
            with open(filename, 'r') as fp:
                info = json.load(fp)
                self.name = info["name"]
                self.phone = info["phone"]
                self.address = info["address"]

        elif self.Type == 'txt':
            print("正在对{filename}进行pickle反序列化...".format(filename=filename))
            with open(filename, 'rb') as fp:
                info = pickle.load(fp)
                self.name = info["name"]
                self.phone = info["phone"]
                self.address = info["address"]

        elif self.Type == 'Tanzq':
            filename = filename.replace(".dat", "")
            print("正在对{filename}进行shelve反序列化...".format(filename=filename))
            with shelve.open(filename) as fp:
                self.name = fp["name"]
                self.phone = fp["phone"]
                self.address = fp["address"]

        print('{name}相关数据反序列化操作完成'.format(name=self.name))

    def createFile(self, extname):
        # 创建对应类型的文件
        FileName = "Data\\" + self.name + extname

        # 创建对应的类型文件，并补充格式并填写
        if extname == '.xlsx':
            file = openpyxl.Workbook()
            sheet = file.active
            sheet.title = self.name
            sheet.cell(1, 1, "姓名")
            sheet.cell(1, 2, '地址')
            sheet.cell(1, 3, '电话')

            if self.name != "none":
                sheet.cell(2, 1, self.name)
            if self.address != "none":
                sheet.cell(2, 2, self.address)
            if self.phone != 'none':
                sheet.cell(2, 3, self.phone)
            file.save(FileName)

        elif extname == '.txt':
            with open(FileName, "w") as fp:
                fp.write("姓名：")
                if self.name != 'none':
                    fp.write(str(self.name))

                fp.write("\n地址：")
                if self.address != 'none':
                    fp.write(str(self.address))

                fp.write("\n电话：")
                if self.phone != 'none':
                    fp.write(str(self.phone))

        elif extname == '.docx':
            doc = Document()
            table = doc.add_table(rows=3, cols=2, style='Table Grid')

            table.cell(0, 0).text = "姓名"
            if self.name != 'none':
                table.rows[0].cells[1].text = self.name

            table.rows[1].cells[0].text = "地址"
            if self.address != 'none':
                table.rows[1].cells[1].text = self.address

            table.rows[2].cells[0].text = "电话"
            if self.phone != 'none':
                table.rows[2].cells[1].text = self.phone

            doc.save(FileName)

        return FileName
